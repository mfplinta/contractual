import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_00

from django.shortcuts import get_object_or_404

from .utils import border_range_colrow, style_range, border_range

from core.models import Configuration
from jobs.models import Job, JobMaterial, JobGroup, JobLabor

COLORS = {
    'header': 'FFF2CB',
    'group_header': 'B6DDE8',
    'material_header': 'DEEBF6',
    'labor_header': 'FBE4D5',
    'totals_header': 'E2EFD9',
}

def build_job_workbook(job_id, show_labor_details=False):
    job = get_object_or_404(Job, pk=job_id)
    groups = JobGroup.objects.filter(job=job).order_by('sort_order', 'id')

    tax_rate = float(job.tax_rate) if job.tax_rate else 0.0
    if not tax_rate:
        settings = Configuration.objects.filter(key='taxRate').first()
        tax_rate = float(settings.value) if settings else 0.0

    output = io.BytesIO()
    workbook = openpyxl.Workbook()
    ws = workbook.active
    assert ws is not None

    # -- General setup --
    ws.title = 'Bill'
    ws.sheet_view.showGridLines = False
    ws.page_setup.fitToPage = True
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    ws.print_options.horizontalCentered = True
    
    # -- Resize columns/rows --
    column_widths = [1.29, 0.67, 5, 4, 38.57, 8.43, 11.43, 1.14, 11.43, 2, 4.86, 9, 0.67]
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # -- Invoice header --
    ws.merge_cells('B2:M2')
    ws['B2'].value = 'Labor & Materials Bill'
    ws['B2'].font = Font(bold=True, size=18)
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B2'].fill = PatternFill(fill_type='solid', start_color=COLORS['header'])
    border_range(ws, 'B2:M2', 'medium')

    # -- Address --
    border_range(ws, 'B4:E7', 'medium')
    style_range(ws, 'B4:E7', fill=PatternFill(fill_type='solid', start_color=COLORS['header']))

    ws['D5'].value = 'Address'
    ws['D5'].font = Font(bold=True)
    ws['D5'].alignment = Alignment(horizontal='right', indent=1)
    ws['D6'].value = 'City'
    ws['D6'].font = Font(bold=True)
    ws['D6'].alignment = Alignment(horizontal='right', indent=1)
    ws['E5'].value = job.client.address if job.client else ''
    ws['E6'].value = job.client.city if job.client else ''

    # -- Materials --

    start_col = 3
    current_row = 9

    cols = {
        'qty': {'col': start_col + 0, 'text': 'Qty', 'align': 'left'},
        'unit': {'col': start_col + 1, 'text': '', 'align': 'left'},
        'item': {'col': start_col + 2, 'text': 'Item', 'align': 'left'},
        'store': {'col': start_col + 3, 'text': 'Store', 'align': 'left'},
        'unit_price': {'col': start_col + 4, 'text': 'Unit Price', 'align': 'right', 'format': FORMAT_NUMBER_00},
        'tax': {'col': start_col + 6, 'text': 'Tax', 'align': 'right', 'format': FORMAT_NUMBER_00},
        'tax_perc': {'col': start_col + 7, 'text': '%', 'align': 'center', 'size': 8},
        'tax_value': {'col': start_col + 8, 'text': f"{tax_rate * 100:.2f}", 'align': 'left', 'size': 8},
        'total': {'col': start_col + 9, 'text': 'Total', 'align': 'right', 'format': FORMAT_NUMBER_00},
    }

    for group in groups:
        group_materials = JobMaterial.objects.filter(group=group).select_related(
            'variant', 'variant__material', 'variant__unit', 'store',
        ).order_by('sort_order', 'id')

        # Group header
        if group.name:
            ws.merge_cells(f'B{current_row}:M{current_row}')
            style_range(ws, f'B{current_row}:M{current_row}', fill=PatternFill(fill_type='solid', start_color=COLORS['group_header']), alignment=Alignment(horizontal='center', vertical='center'), font=Font(bold=True, size=12))
            ws[f'B{current_row}'].value = group.name
            border_range(ws, f'B{current_row}:M{current_row}', 'medium')
            ws.row_dimensions[current_row].height = 25
            current_row += 1
            ws.row_dimensions[current_row].height = 6
            current_row += 1

        # Materials title
        start_border_row = current_row
        ws.merge_cells(f'B{current_row}:M{current_row}')
        style_range(ws, f'B{current_row}:M{current_row}',
                    fill=PatternFill(fill_type='solid', start_color=COLORS['material_header']),
                    alignment=Alignment(horizontal='center'))
        ws[f'B{current_row}'].value = 'Materials'

        # Column headers
        for col in cols.values():
            cell = ws.cell(row=current_row + 1, column=col['col'])
            cell.value = col['text']
            cell.font = Font(size=col.get('size', 11), bold=True)
            cell.alignment = Alignment(horizontal=col['align'])

        # Data rows
        data_start_row = current_row + 2
        for row_idx, jm in enumerate(group_materials, start=data_start_row):
            for col in cols.values():
                cell = ws.cell(row=row_idx, column=col['col'])
                if 'align' in col:
                    cell.alignment = Alignment(horizontal=col['align'])
                if 'format' in col:
                    cell.number_format = col['format']
            
            ws.cell(row=row_idx, column=cols['qty']['col'], value=float(jm.quantity))
            ws.cell(
                row=row_idx,
                column=cols['unit']['col'],
                value=(jm.variant.unit.shorthand or jm.variant.unit.name) if jm.variant and jm.variant.unit else '',
            )
            ws.cell(row=row_idx, column=cols['item']['col'], value=(f"{jm.variant.name} - " if jm.variant and jm.variant.name else "") + (jm.variant.material.description if jm.variant and jm.variant.material else ""))
            ws.cell(row=row_idx, column=cols['store']['col'], value=jm.store.name if jm.store else '')
            ws.cell(row=row_idx, column=cols['unit_price']['col'], value='?' if jm.ignored else float(jm.unit_price))
            ws.cell(row=row_idx, column=cols['tax']['col'], value='?' if jm.ignored else float(jm.tax))
            ws.cell(row=row_idx, column=cols['total']['col'], value='?' if jm.ignored else float(jm.total_price))
            if jm.ignored:
                style_range(ws, f"G{row_idx}:M{row_idx}", font=Font(color='FF0000'))

        num_materials = group_materials.count()

        # Group subtotal row
        subtotal_row = data_start_row + num_materials
        ws.merge_cells(start_row=subtotal_row, start_column=cols['qty']['col'] - 1, end_row=subtotal_row, end_column=cols['tax']['col'] - 1)
        ws.cell(row=subtotal_row, column=cols['qty']['col'] - 1, value='Materials Total')
        ws.cell(row=subtotal_row, column=cols['qty']['col'] - 1).font = Font(bold=True)
        ws.cell(row=subtotal_row, column=cols['qty']['col'] - 1).alignment = Alignment(horizontal='right')
        ws.cell(row=subtotal_row, column=cols['tax']['col'], value=float(group.tax_total))
        ws.cell(row=subtotal_row, column=cols['tax']['col']).number_format = FORMAT_NUMBER_00
        ws.cell(row=subtotal_row, column=cols['tax']['col']).font = Font(bold=True)
        ws.cell(row=subtotal_row, column=cols['total']['col'], value=float(group.total))
        ws.cell(row=subtotal_row, column=cols['total']['col']).number_format = FORMAT_NUMBER_00
        ws.cell(row=subtotal_row, column=cols['total']['col']).font = Font(bold=True)
        
        current_row = subtotal_row + 1
        section_end_row = subtotal_row

        # -- Labor section --
        labor_items = JobLabor.objects.filter(group=group).order_by('id')

        if labor_items.exists():
            ws.merge_cells(f'B{current_row}:M{current_row}')
            style_range(ws, f'B{current_row}:M{current_row}',
                        fill=PatternFill(fill_type='solid', start_color=COLORS['labor_header']),
                        alignment=Alignment(horizontal='center'))
            ws[f'B{current_row}'].value = 'Labor'

            labor_header_row = current_row + 1

            if group.labor_time_total > 0:
                ws.cell(row=labor_header_row, column=cols['tax']['col'], value='Hours')
                ws.cell(row=labor_header_row, column=cols['tax']['col']).font = Font(size=11, bold=True)
                ws.cell(row=labor_header_row, column=cols['tax']['col']).alignment = Alignment(horizontal='right')
            if group.labor_cost_total > 0:
                ws.cell(row=labor_header_row, column=cols['total']['col'], value='Total')
                ws.cell(row=labor_header_row, column=cols['total']['col']).font = Font(size=11, bold=True)
                ws.cell(row=labor_header_row, column=cols['total']['col']).alignment = Alignment(horizontal='right')

            if show_labor_details:
                # Labor column headers
                ws.cell(row=labor_header_row, column=cols['item']['col'], value='Item')
                ws.cell(row=labor_header_row, column=cols['item']['col']).font = Font(size=11, bold=True)

                # Labor data rows
                labor_data_start = labor_header_row + 1
                for li_idx, li in enumerate(labor_items, start=labor_data_start):
                    ws.cell(row=li_idx, column=cols['item']['col'], value=li.description)
                    if li.time > 0:
                        ws.cell(row=li_idx, column=cols['tax']['col'], value=li.time)
                        ws.cell(row=li_idx, column=cols['tax']['col']).number_format = FORMAT_NUMBER_00
                    if li.cost > 0:
                        ws.cell(row=li_idx, column=cols['total']['col'], value=li.cost)
                        ws.cell(row=li_idx, column=cols['total']['col']).number_format = FORMAT_NUMBER_00
                    ws.cell(row=li_idx, column=cols['total']['col']).alignment = Alignment(horizontal='right')
                
                current_row = labor_data_start + labor_items.count()
            else:
                current_row += 2

            # Labor total line
            ws.merge_cells(start_row=current_row, start_column=cols['qty']['col'] - 1, end_row=current_row, end_column=cols['tax']['col'] - 1)
            ws.cell(row=current_row, column=cols['qty']['col'] - 1, value='Labor Total')
            ws.cell(row=current_row, column=cols['qty']['col'] - 1).font = Font(bold=True)
            ws.cell(row=current_row, column=cols['qty']['col'] - 1).alignment = Alignment(horizontal='right')
            if group.labor_time_total > 0:
                ws.cell(row=current_row, column=cols['tax']['col'], value=group.labor_time_total)
                ws.cell(row=current_row, column=cols['tax']['col']).number_format = FORMAT_NUMBER_00
                ws.cell(row=current_row, column=cols['tax']['col']).font = Font(bold=True)
            if group.labor_cost_total > 0:
                ws.cell(row=current_row, column=cols['total']['col'], value=group.labor_cost_total)
                ws.cell(row=current_row, column=cols['total']['col']).number_format = FORMAT_NUMBER_00
                ws.cell(row=current_row, column=cols['total']['col']).font = Font(bold=True)

            section_end_row = current_row

            current_row += 1

        border_range_colrow(
            ws,
            min_col=start_col - 1,
            min_row=start_border_row,
            max_col=start_col + 10,
            max_row=section_end_row,
            border_style='medium',
        )

        ws.row_dimensions[current_row].height = 6
        current_row += 1

        ws.merge_cells(f'B{current_row}:M{current_row}')
        style_range(ws, f'B{current_row}:M{current_row}', fill=PatternFill(fill_type='solid', start_color=COLORS['totals_header']), alignment=Alignment(horizontal='center'))
        ws[f'B{current_row}'].value = 'Totals'

        current_row += 1

        # -- Group totals --
        ws.cell(row=current_row, column=cols['qty']['col'], value='Material:')
        ws.cell(row=current_row, column=cols['total']['col'], value=float(group.total))
        ws.cell(row=current_row, column=cols['total']['col']).number_format = FORMAT_NUMBER_00

        current_row += 1
        ws.cell(row=current_row, column=cols['qty']['col'], value='Labor:')
        if group.labor_cost_total <= 0 and group.labor_time_total > 0:
            ws.cell(row=current_row, column=cols['total']['col'], value=f'{group.labor_time_total:.2f} hours')
        else:
            ws.cell(row=current_row, column=cols['total']['col'], value=group.labor_cost_total)
            ws.cell(row=current_row, column=cols['total']['col']).number_format = FORMAT_NUMBER_00
        ws.cell(row=current_row, column=cols['total']['col']).alignment = Alignment(horizontal='right')

        current_row += 1
        ws.cell(row=current_row, column=cols['qty']['col'], value='Invoice total:')
        ws.cell(row=current_row, column=cols['qty']['col']).font = Font(bold=True)
        ws.cell(row=current_row, column=cols['total']['col'], value=float(group.total + group.labor_cost_total))
        ws.cell(row=current_row, column=cols['total']['col']).number_format = FORMAT_NUMBER_00
        ws.cell(row=current_row, column=cols['total']['col']).font = Font(bold=True)

        current_row += 1
        ws.cell(row=current_row, column=cols['qty']['col'], value='Status:')

        border_range_colrow(ws, min_col=start_col - 1, min_row=current_row - 4,
                           max_col=start_col + 10, max_row=current_row,
                           border_style='medium')

        current_row += 2

    workbook.save(output)
    output.seek(0)

    client_name = job.client.name if job.client else 'Unknown'
    filename = f"{client_name} - {job.description}.xlsx".replace('/', '-')

    return output, filename
